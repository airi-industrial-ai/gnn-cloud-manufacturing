import numpy as np
import pandas as pd
from tqdm.auto import tqdm
from openpyxl import load_workbook
from joblib import Parallel, delayed

def generate_tariff_matrix(num_companies, distance_matrix, productivity,
                           cost_operations, variability=0.1):
    num_cities = distance_matrix.shape[0]
    tariff_matrix = np.zeros((num_companies, num_cities, num_cities))

    # if it's unprofitale to produce something in the city than it's more expensive logistics
    baseline = np.mean(cost_operations, axis=0).reshape(-1,1)
    # higher productivity makes price lower
    operational_tariffs = (distance_matrix+baseline*0.1) * (1 - productivity.reshape(-1, 1))
    random_factors = 1 + np.random.randn(num_companies, num_cities, num_cities) * variability
    tariff_matrix = operational_tariffs * random_factors[:, :, :]
    for company in range(num_companies):
        np.fill_diagonal(tariff_matrix[company], 0)
    return tariff_matrix

def read_fatahi_dataset(path_to_file):
    np.random.seed(42)
    """
    Fatahi Valilai, Omid. “Dataset for Logistics and Manufacturing
    Service Composition”. 17 Mar. 2021. Web. 9 June 2023.
    """
    workbook = load_workbook(filename=path_to_file, read_only=True)
    sheet_names = workbook.sheetnames

    res = []
    def process_sheet(sheet_name):
        return _read_sheet(path_to_file, sheet_name)
    res = Parallel(n_jobs=-1)(
        delayed(process_sheet)(sheet_name) for sheet_name in tqdm(sheet_names)
    )
    return res

def _read_sheet(path_to_file, sheet_name):
    n_services = 10
    n_operations, n_suboperations, n_cities, _ = list(
        map(int, "-".join(sheet_name.split(",")).split("-"))
    )
    operations = np.zeros((n_suboperations, n_operations))
    dist = np.zeros((n_cities, n_cities))
    time_cost = np.zeros((n_suboperations, n_cities))
    op_cost = np.zeros((n_suboperations, n_cities))
    productivity = np.zeros((n_cities))
    transportation_cost = np.zeros((n_services))

    operations[:, :] = pd.read_excel(
        path_to_file,
        sheet_name=sheet_name,
        header=None,
        usecols=range(1, n_operations + 1),
        skiprows=5,
        nrows=n_suboperations,
    )
    dist[:, :] = pd.read_excel(
        path_to_file,
        sheet_name=sheet_name,
        header=None,
        usecols=range(1, n_cities + 1),
        skiprows=5 * 2 + n_suboperations - 1,
        nrows=n_cities,
    )
    time_cost[:, :] = pd.read_excel(
        path_to_file,
        sheet_name=sheet_name,
        header=None,
        usecols=range(1, n_cities + 1),
        skiprows=5 * 3 + n_suboperations + n_cities - 1 * 2,
        nrows=n_suboperations,
    )
    time_cost[np.isinf(time_cost)] = 99
    op_cost[:, :] = pd.read_excel(
        path_to_file,
        sheet_name=sheet_name,
        header=None,
        usecols=range(1, n_cities + 1),
        skiprows=5 * 4 + n_suboperations + n_cities + n_suboperations - 1 * 3,
        nrows=n_suboperations,
    )
    op_cost[np.isinf(op_cost)] = 999
    productivity[:] = pd.read_excel(
        path_to_file,
        sheet_name=sheet_name,
        header=None,
        usecols=range(n_cities),
        skiprows=5 * 5
        + n_suboperations
        + n_cities
        + n_suboperations
        + n_suboperations
        - 1 * 4,
        nrows=1,
    )
    transportation_cost = generate_tariff_matrix(10, dist, productivity, op_cost)
    return {
        "name": sheet_name,
        "n_operations": n_operations,
        "n_suboperations": n_suboperations,
        "n_cities": n_cities,
        "n_services": n_services,
        "operations": operations,
        "dist": dist,
        "time_cost": time_cost,
        "op_cost": op_cost,
        "productivity": productivity.reshape(1,-1),
        "transportation_cost": transportation_cost,
    }
