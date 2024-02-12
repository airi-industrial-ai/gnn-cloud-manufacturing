import torch
import dgl
import numpy as np
from sklearn.preprocessing import OneHotEncoder
import torch.nn.functional as F
import openpyxl
from cloudmanufacturing.mip_solver import mip_solve
import pickle
from cloudmanufacturing.data import solve_excel

ss_type = ('s', 'ss', 's')
os_type = ('o', 'os', 's')
so_type = ('s', 'so', 'o')

def calculate_optimal(train_pth, logger, dataset, os_type, ss_type, stop_count=100):

	filepath = train_pth[:-5] + '_solved.xlsx'
	solve_pth = train_pth[:-5] + '_solve.pickle'
	
	wb = openpyxl.Workbook()
	wb.save(filepath)
	DGList =[]
	counter = 0

	for problem in dataset:
		if counter <= stop_count:
			delta, train_gamma, status, value = mip_solve(problem)

			if status.name == "OPTIMAL":
				try:
					
					example_graph = dglgraph_fixed(problem, train_gamma)
					example_target = example_graph.edata['target'][os_type]
					example_graph.edata['feat'][os_type][:, 0] /= 10
					example_graph.edata['feat'][ss_type][:] /= 100
					DGList.append(example_graph)
					with open(solve_pth, 'wb') as f:
						pickle.dump(DGList, f)
			
					logger.info(f"Solved as OPTIMAL and saved in dgl format task with parameters {problem['name']}")
					sheet_name = problem['name']
					city_to_op_solution = []
					otc=[]
					for o, t, c in zip(*np.where(train_gamma == 1)):
						city_to_op_solution.append((f'city {c}', f'{t}_{o}'))
						otc.append([o,t,c])

					shape_g = np.shape(train_gamma)
					solve_excel(filepath, sheet_name, city_to_op_solution, otc, shape_g)
					logger.info(f"Saved as excel task with parameters {problem['name']}")
					counter+=1

				except:
					logger.exception('Empty operation included!')
					pass
					
			source_workbook = openpyxl.load_workbook(train_pth[:-5] + '_solved.xlsx')
			source_sheet_names = source_workbook.sheetnames[1:]
			train_workbook = openpyxl.load_workbook(train_pth)
			new_workbook = openpyxl.Workbook()

			# Копируем листы из исходного файла в новый файл
			for sheet_name in source_sheet_names:
				source_sheet = train_workbook[sheet_name]
				new_sheet = new_workbook.create_sheet(title=sheet_name)
					
				for row in source_sheet.iter_rows(values_only=True):
					new_sheet.append(row)

			# Сохраняем новый Excel-файл
			new_workbook.save(train_pth[:-5] + '_OPTIMAL.xlsx')
			with open(train_pth[:-5] + '_sheet_names.pickle', 'wb') as s:
				pickle.dump(source_sheet_names, s)
			

def dglgraph_fixed(problem, gamma, oper_max=20):
	g = dglgraph(problem, gamma)
	ncolumns = g.ndata['feat']['o'].shape[1]
	g.ndata['feat'] = {'o': F.pad(g.ndata['feat']['o'], [0, oper_max - ncolumns])}
	return g


def dglgraph(problem, gamma):
	n_tasks = problem['n_tasks']
	n_operations = problem['n_operations']
	operation = problem['operation']
	dist = problem['dist']
	time_cost = problem['time_cost']
	op_cost = problem['op_cost']
	productivity = problem['productivity']

	operation_index = []
	for i in range(n_tasks):
		for j in range(n_operations):
			if operation[j, i] == 1:
				operation_index.append((i, j))
	operation_index.append([n_tasks // 2, -1])
	operation_index.append([n_tasks // 2, n_operations])
	operation_index = np.array(operation_index)
	
	adj_operation = np.zeros((operation_index.shape[0], operation_index.shape[0]))
	for i in range(n_tasks):
		col_i = operation[:, i]
		path = np.where(col_i > 0)[0]
		for j in range(len(path) - 1):
			u = operation_index.tolist().index([i, path[j]])
			v = operation_index.tolist().index([i, path[j+1]])
			adj_operation[u, v] = 1
		adj_operation[-2, operation_index.tolist().index([i, path[0]])] = 1
		adj_operation[operation_index.tolist().index([i, path[-1]]), -1] = 1

	full_time_cost = np.tile(time_cost, (n_tasks, 1))
	full_time_cost = full_time_cost[operation.T.reshape(-1).astype(bool)]

	full_op_cost = np.tile(op_cost, (n_tasks, 1))
	full_op_cost = full_op_cost[operation.T.reshape(-1).astype(bool)]

	target = []
	for full_o, c in zip(*np.where(full_op_cost < 999)):
		t, o = operation_index[full_o]
		target.append(gamma[o, t, c])

	graph_data = {
		ss_type: np.where(dist > 0),
		os_type: np.where(full_op_cost < 999),
		so_type: np.where(full_op_cost < 999)[::-1],
		('o', 'forward', 'o'): np.where(adj_operation > 0),
		('o', 'backward', 'o'): np.where(adj_operation > 0)[::-1],
	}
	g = dgl.heterograph(graph_data)
	g = dgl.add_self_loop(g, etype='ss')

	op_feat = OneHotEncoder().fit_transform(
		operation_index[g.nodes('o').numpy()][:, [1]]
	).toarray()
	g.ndata['feat'] = {
		'o': torch.FloatTensor(op_feat),
		's': torch.FloatTensor(productivity[:, None])
	}
	g.ndata['operation_index'] = {
		'o': torch.LongTensor(operation_index),
	}
	u_idx, v_idx = g.edges(etype='os')
	serves_feat = np.array([
		full_op_cost[u_idx, v_idx],
		full_time_cost[u_idx, v_idx],
	])
	g.edata['feat'] = {
		'os': torch.FloatTensor(serves_feat.T),
		'ss': torch.FloatTensor(dist[g.edges(etype='ss')][:, None]),
	}
	g.edata['target'] = {
		'os': torch.FloatTensor(target)[:, None],
	}
	return g


def graph_gamma(graph, problem):
	target_mask = graph.edata['target'][os_type][:, 0] == 1
	u, v = graph.edges(etype=os_type, )
	u, v = u[target_mask], v[target_mask]
	u = graph.ndata['operation_index']['o'][u]
	gamma = np.zeros((problem['n_operations'], problem['n_tasks'], problem['n_cities']))
	for i in range(len(u)):
		operation, task, city = u[i, 1], u[i, 0], v[i]
		gamma[operation, task, city] = 1
	return gamma
